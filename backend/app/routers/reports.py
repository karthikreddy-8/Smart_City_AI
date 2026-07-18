import io
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from app.database import get_db
from app.routers.analytics import get_base_dataframe
from app.auth import get_current_user, User

router = APIRouter(prefix="/reports", tags=["Report Generation"])

@router.get("/csv")
def download_csv_report(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    df = get_base_dataframe(db)
    if df.empty:
        raise HTTPException(status_code=404, detail="No traffic data available to generate report")

    stream = io.StringIO()
    df.to_csv(stream, index=False)

    response = StreamingResponse(iter([stream.getvalue()]), media_type="text/csv")
    response.headers["Content-Disposition"] = "attachment; filename=traffic_report.csv"
    return response

@router.get("/excel")
def download_excel_report(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    df = get_base_dataframe(db)
    if df.empty:
        raise HTTPException(status_code=404, detail="No traffic data available to generate report")

    wb = Workbook()
    ws = wb.active
    ws.title = "SmartCity AI Analytics"

    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")

    headers = list(df.columns)
    ws.append(headers)

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    for _, row in df.iterrows():
        ws.append(list(row))

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max(max_len + 3, 10)

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    response = StreamingResponse(stream, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response.headers["Content-Disposition"] = "attachment; filename=traffic_analytics_report.xlsx"
    return response

@router.get("/pdf")
def download_pdf_report(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    df = get_base_dataframe(db)
    if df.empty:
        raise HTTPException(status_code=404, detail="No traffic data available to generate report")

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)
    story = []

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=24,
        textColor=colors.HexColor('#0f172a'),
        spaceAfter=15
    )

    body_style = ParagraphStyle(
        'DocBody',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=10,
        textColor=colors.HexColor('#334155'),
        spaceAfter=10
    )

    story.append(Paragraph("SmartCity AI – Traffic Analytics Executive Report", title_style))
    story.append(Paragraph("Generated automatically by SmartCity Traffic Congestion Analytics & Intelligent Road Optimization System.", body_style))
    story.append(Spacer(1, 15))

    stats_data = [
        ["Metric", "Value"],
        ["Total Recorded Observations", str(len(df))],
        ["Average Observed Speed", f"{round(df['Average Speed'].mean(), 1)} km/h"],
        ["Average Vehicle Count", f"{round(df['Vehicle Count'].mean(), 0)} vehicles"],
        ["Total Recorded Accidents", str(int(df['Accident Count'].sum()))],
        ["Highest Congested Road", str(df.groupby("Road Name")["Vehicle Count"].mean().idxmax())]
    ]

    table = Table(stats_data, colWidths=[200, 300])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1e293b')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('BOTTOMPADDING', (0,0), (-1,0), 8),
        ('BACKGROUND', (0,1), (-1,-1), colors.HexColor('#f8fafc')),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#cbd5e1')),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 11),
        ('FONTNAME', (0,1), (-1,-1), 'Helvetica'),
        ('FONTSIZE', (0,1), (-1,-1), 10),
    ]))
    story.append(table)
    story.append(Spacer(1, 20))

    story.append(Paragraph("Detailed Segment Analytics (Top 10 Samples)", ParagraphStyle('SubTitle', parent=styles['Heading2'], fontSize=14, spaceAfter=8)))

    road_data = [["Road Name", "Road Type", "Avg Vehicles", "Avg Speed (km/h)", "Congestion"]]
    sample_df = df.head(10)
    for _, row in sample_df.iterrows():
        road_data.append([
            str(row["Road Name"]),
            str(row["Road Type"]),
            str(int(row["Vehicle Count"])),
            str(round(row["Average Speed"], 1)),
            str(row["Congestion Level"])
        ])

    road_table = Table(road_data, colWidths=[150, 100, 90, 100, 80])
    road_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#0f766e')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('BOTTOMPADDING', (0,0), (-1,0), 6),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#e2e8f0')),
        ('BACKGROUND', (0,1), (-1,-1), colors.HexColor('#fafafa')),
    ]))
    story.append(road_table)

    doc.build(story)
    buffer.seek(0)

    response = StreamingResponse(buffer, media_type="application/pdf")
    response.headers["Content-Disposition"] = "attachment; filename=traffic_executive_report.pdf"
    return response
